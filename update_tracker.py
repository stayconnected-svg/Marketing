"""
Update the Bryan Outreach Tracker with email campaign tracking columns.
Adds sequence tracking, deliverability metrics, and campaign analytics.
"""
import os

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
TRACKER_PATH = os.path.join(OUTPUT_DIR, "Bryan_Outreach_Tracker.xlsx")

EMAIL_CAMPAIGN_HEADERS = [
    "Email Verified",
    "Email Source",
    "Domain",
    "Sequence Name",
    "Sequence Step",
    "Email 1 Sent",
    "Email 1 Opened",
    "Email 1 Replied",
    "Email 2 Sent",
    "Email 2 Opened",
    "Email 2 Replied",
    "Email 3 Sent",
    "Email 3 Opened",
    "Email 3 Replied",
    "Email 4 Sent",
    "Email 4 Opened",
    "Email 4 Replied",
    "Email 5 Sent",
    "Email 5 Opened",
    "Email 5 Replied",
    "Total Opens",
    "Total Replies",
    "Bounced (Y/N)",
    "Bounce Type",
    "Unsubscribed (Y/N)",
    "Marked Spam (Y/N)",
    "Sending Mailbox",
    "Sending Domain",
    "Warm Handoff Sent",
    "Handoff to Bryan Date",
]

wb = load_workbook(TRACKER_PATH)
ws = wb["Prospect Tracker"]

header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
campaign_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

existing_cols = ws.max_column
start_col = existing_cols + 1

for i, header in enumerate(EMAIL_CAMPAIGN_HEADERS):
    col = start_col + i
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = campaign_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col)].width = 16

# Update auto-filter to include new columns
ws.auto_filter.ref = f"A1:{get_column_letter(start_col + len(EMAIL_CAMPAIGN_HEADERS) - 1)}1"

# Update Dashboard with email campaign metrics
dash = wb["Dashboard"]

campaign_metrics_start = 15
campaign_metrics = [
    ("", ""),
    ("--- EMAIL CAMPAIGN METRICS ---", ""),
    ("Total Emails Sent", "=SUM('Prospect Tracker'!{c1}:{c1})".format(
        c1=get_column_letter(start_col + 5))),  # placeholder
    ("Emails Opened (unique prospects)", ""),
    ("Emails Replied (unique prospects)", ""),
    ("Bounced", ""),
    ("Unsubscribed", ""),
    ("", ""),
    ("--- DELIVERABILITY ---", ""),
    ("Open Rate", ""),
    ("Reply Rate", ""),
    ("Bounce Rate", ""),
    ("Spam Complaint Rate", ""),
    ("", ""),
    ("--- SEQUENCE PERFORMANCE ---", ""),
    ("Email 1 Open Rate", ""),
    ("Email 2 Open Rate", ""),
    ("Email 3 Open Rate", ""),
    ("Email 4 Open Rate", ""),
    ("Email 5 Open Rate", ""),
    ("", ""),
    ("--- PIPELINE ---", ""),
    ("Warm Handoffs Sent", ""),
    ("Handed Off to Bryan", ""),
    ("Meetings from Email", ""),
]

bold_font = Font(bold=True)
section_font = Font(bold=True, color="2E7D32")

for i, (metric, value) in enumerate(campaign_metrics):
    row = campaign_metrics_start + i
    cell_a = dash.cell(row=row, column=1, value=metric)
    cell_b = dash.cell(row=row, column=2, value=value)
    if metric.startswith("---"):
        cell_a.font = section_font
    else:
        cell_a.font = bold_font

# Add reference data for email tracking
ref = wb["Reference"]
ref.cell(row=1, column=7, value="Email Statuses").font = Font(bold=True)
email_statuses = [
    "verified", "pattern_match", "catchall", "no_mx",
    "invalid", "unknown", "error",
]
for i, status in enumerate(email_statuses, 2):
    ref.cell(row=i, column=7, value=status)

ref.cell(row=1, column=9, value="Bounce Types").font = Font(bold=True)
bounce_types = ["hard", "soft", "invalid_address", "mailbox_full", "domain_error"]
for i, bt in enumerate(bounce_types, 2):
    ref.cell(row=i, column=9, value=bt)

ref.cell(row=1, column=11, value="Sequence Names").font = Font(bold=True)
sequences = [
    "Simpl ABI - Dentist Sequence",
    "Simpl ABI - Doctor Sequence",
    "Simpl ABI - Lawyer Sequence",
    "Bryan Direct - Pre-Approved",
]
for i, seq in enumerate(sequences, 2):
    ref.cell(row=i, column=11, value=seq)

ref.column_dimensions["G"].width = 18
ref.column_dimensions["I"].width = 18
ref.column_dimensions["K"].width = 35

wb.save(TRACKER_PATH)
print(f"Updated tracker: {TRACKER_PATH}")
print(f"  Added {len(EMAIL_CAMPAIGN_HEADERS)} email campaign columns (green headers)")
print(f"  Added email campaign metrics to Dashboard")
print(f"  Added email reference data to Reference sheet")
