"""
Create a CRM-ready Excel tracking sheet for Bryan McInerney's outreach.
Includes columns for prospect data, outreach tracking, and follow-up management.
"""
import csv
import os

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    USE_XLSX = True
except ImportError:
    USE_XLSX = False

HEADERS = [
    "Priority",
    "Industry",
    "Full Name",
    "Title/Credential",
    "Business Name",
    "Specialty/Practice Area",
    "Address",
    "City",
    "State",
    "ZIP",
    "Phone",
    "Email",
    "LinkedIn URL",
    "Website",
    "Source",
    "Date Added",
    "Outreach 1 Date",
    "Outreach 1 Channel",
    "Outreach 1 Template",
    "Outreach 2 Date",
    "Outreach 2 Channel",
    "Outreach 2 Template",
    "Outreach 3 Date",
    "Outreach 3 Channel",
    "Outreach 3 Template",
    "Response (Y/N)",
    "Response Date",
    "Response Type",
    "Meeting Scheduled (Y/N)",
    "Meeting Date",
    "Status",
    "Next Step",
    "Notes",
]

STATUS_OPTIONS = [
    "New",
    "Contacted",
    "Responded",
    "Meeting Set",
    "Meeting Complete",
    "Proposal Sent",
    "Won",
    "Lost",
    "Not Interested",
    "Follow Up Later",
]

CHANNEL_OPTIONS = [
    "Email - Cold Intro",
    "Email - Whitepaper Offer",
    "Email - Event Invite",
    "Email - Follow Up",
    "LinkedIn - Connection Request",
    "LinkedIn - Message",
    "Phone Call",
    "Direct Mail",
    "In Person",
    "Referral",
]


def create_xlsx():
    """Create a formatted Excel tracking sheet."""
    wb = Workbook()

    # Main tracking sheet
    ws = wb.active
    ws.title = "Prospect Tracker"

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    col_widths = {
        "Priority": 10, "Industry": 12, "Full Name": 22, "Title/Credential": 15,
        "Business Name": 28, "Specialty/Practice Area": 22, "Address": 30,
        "City": 15, "State": 8, "ZIP": 10, "Phone": 16, "Email": 28,
        "LinkedIn URL": 30, "Website": 25, "Source": 15, "Date Added": 12,
        "Outreach 1 Date": 14, "Outreach 1 Channel": 20, "Outreach 1 Template": 22,
        "Outreach 2 Date": 14, "Outreach 2 Channel": 20, "Outreach 2 Template": 22,
        "Outreach 3 Date": 14, "Outreach 3 Channel": 20, "Outreach 3 Template": 22,
        "Response (Y/N)": 12, "Response Date": 14, "Response Type": 18,
        "Meeting Scheduled (Y/N)": 14, "Meeting Date": 14, "Status": 16,
        "Next Step": 25, "Notes": 35,
    }
    for col_idx, header in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 15)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    ws.freeze_panes = "A2"

    # Dashboard sheet
    dash = wb.create_sheet("Dashboard")
    dash_headers = ["Metric", "Value"]
    for col_idx, h in enumerate(dash_headers, 1):
        cell = dash.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    metrics = [
        ("Total Prospects", '=COUNTA(\'Prospect Tracker\'!C:C)-1'),
        ("Dentists", '=COUNTIF(\'Prospect Tracker\'!B:B,"Dentist")'),
        ("Doctors", '=COUNTIF(\'Prospect Tracker\'!B:B,"Physician")'),
        ("Lawyers", '=COUNTIF(\'Prospect Tracker\'!B:B,"Lawyer")'),
        ("", ""),
        ("Contacted", '=COUNTIF(\'Prospect Tracker\'!AG:AG,"Contacted")+COUNTIF(\'Prospect Tracker\'!AG:AG,"Responded")+COUNTIF(\'Prospect Tracker\'!AG:AG,"Meeting Set")+COUNTIF(\'Prospect Tracker\'!AG:AG,"Meeting Complete")+COUNTIF(\'Prospect Tracker\'!AG:AG,"Proposal Sent")+COUNTIF(\'Prospect Tracker\'!AG:AG,"Won")'),
        ("Responded", '=COUNTIF(\'Prospect Tracker\'!AA:AA,"Y")'),
        ("Meetings Scheduled", '=COUNTIF(\'Prospect Tracker\'!AD:AD,"Y")'),
        ("Won", '=COUNTIF(\'Prospect Tracker\'!AG:AG,"Won")'),
        ("", ""),
        ("Response Rate", '=IF(B6>0,B7/B6,"N/A")'),
        ("Meeting Rate", '=IF(B7>0,B8/B7,"N/A")'),
    ]
    for row_idx, (metric, value) in enumerate(metrics, 2):
        dash.cell(row=row_idx, column=1, value=metric).font = Font(bold=True)
        dash.cell(row=row_idx, column=2, value=value)

    dash.column_dimensions["A"].width = 25
    dash.column_dimensions["B"].width = 15

    # Reference sheet
    ref = wb.create_sheet("Reference")
    ref.cell(row=1, column=1, value="Status Options").font = Font(bold=True)
    for i, status in enumerate(STATUS_OPTIONS, 2):
        ref.cell(row=i, column=1, value=status)

    ref.cell(row=1, column=3, value="Channel Options").font = Font(bold=True)
    for i, channel in enumerate(CHANNEL_OPTIONS, 2):
        ref.cell(row=i, column=3, value=channel)

    ref.cell(row=1, column=5, value="Industries").font = Font(bold=True)
    for i, ind in enumerate(["Dentist", "Physician", "Lawyer"], 2):
        ref.cell(row=i, column=5, value=ind)

    ref.column_dimensions["A"].width = 20
    ref.column_dimensions["C"].width = 30
    ref.column_dimensions["E"].width = 15

    output = os.path.join(OUTPUT_DIR, "Bryan_Outreach_Tracker.xlsx")
    wb.save(output)
    print(f"Created Excel tracker: {output}")
    return output


def create_csv_fallback():
    """Create a CSV tracking sheet if openpyxl is not available."""
    output = os.path.join(OUTPUT_DIR, "Bryan_Outreach_Tracker.csv")
    with open(output, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
    print(f"Created CSV tracker: {output}")
    return output


if __name__ == "__main__":
    if USE_XLSX:
        create_xlsx()
    else:
        print("openpyxl not found, trying to install...")
        import subprocess
        subprocess.run(["pip", "install", "openpyxl"], check=True)
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        create_xlsx()
